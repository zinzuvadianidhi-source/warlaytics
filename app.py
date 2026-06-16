import os
from datetime import datetime
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, session, flash
from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
app.secret_key = "change-this-secret-key"

UPLOAD_FOLDER = os.path.join("static", "uploads")
EXCEL_FILE = "database.xlsx"
ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "webp"}
PHOTO_TYPES = ["Front", "Back", "Left", "Right", "Top", "Bottom", "Tag"]

# Demo login. Change these for your real use.
USERS = {
    "admin": {"password": "admin123", "role": "admin"},
    "user": {"password": "user123", "role": "user"},
}

app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Photos"
        ws.append(["ID", "Product Name", "Photo Type", "File Name", "Uploaded By", "Created At"])
        wb.save(EXCEL_FILE)


def get_rows():
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            rows.append({
                "id": row[0],
                "product_name": row[1],
                "photo_type": row[2],
                "file_name": row[3],
                "uploaded_by": row[4],
                "created_at": row[5],
            })
    return rows


def add_row(product_name, photo_type, file_name, uploaded_by):
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    next_id = 1
    ids = [cell.value for cell in ws["A"][1:] if isinstance(cell.value, int)]
    if ids:
        next_id = max(ids) + 1
    ws.append([next_id, product_name, photo_type, file_name, uploaded_by, datetime.now().strftime("%Y-%m-%d %H:%M")])
    wb.save(EXCEL_FILE)


def update_row(photo_id, product_name, photo_type, file_name=None):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 1).value == photo_id:
            ws.cell(row, 2).value = product_name
            ws.cell(row, 3).value = photo_type
            if file_name:
                old_file = ws.cell(row, 4).value
                if old_file:
                    old_path = os.path.join(app.config["UPLOAD_FOLDER"], old_file)
                    if os.path.exists(old_path):
                        os.remove(old_path)
                ws.cell(row, 4).value = file_name
            break
    wb.save(EXCEL_FILE)


def delete_row(photo_id):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 1).value == photo_id:
            file_name = ws.cell(row, 4).value
            if file_name:
                path = os.path.join(app.config["UPLOAD_FOLDER"], file_name)
                if os.path.exists(path):
                    os.remove(path)
            ws.delete_rows(row)
            break
    wb.save(EXCEL_FILE)


def login_required(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        if "username" not in session:
            return redirect(url_for("login"))
        return func(*args, **kwargs)
    return wrapper


@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        user = USERS.get(username)
        if user and user["password"] == password:
            session["username"] = username
            session["role"] = user["role"]
            return redirect(url_for("dashboard"))
        flash("Wrong username or password")
    return render_template("login.html")


@app.route("/dashboard")
@login_required
def dashboard():
    rows = get_rows()
    return render_template("dashboard.html", rows=rows, username=session["username"], role=session["role"])


@app.route("/add", methods=["GET", "POST"])
@login_required
def add_photo():
    if request.method == "POST":
        product_name = request.form.get("product_name", "").strip()
        photo_type = request.form.get("photo_type")
        file = request.files.get("photo")
        if not product_name or photo_type not in PHOTO_TYPES or not file or file.filename == "":
            flash("Please fill all fields")
            return redirect(url_for("add_photo"))
        if not allowed_file(file.filename):
            flash("Only PNG, JPG, JPEG, WEBP images are allowed")
            return redirect(url_for("add_photo"))
        filename = datetime.now().strftime("%Y%m%d%H%M%S_") + secure_filename(file.filename)
        file.save(os.path.join(app.config["UPLOAD_FOLDER"], filename))
        add_row(product_name, photo_type, filename, session["username"])
        return redirect(url_for("dashboard"))
    return render_template("add.html", photo_types=PHOTO_TYPES)


@app.route("/edit/<int:photo_id>", methods=["GET", "POST"])
@login_required
def edit_photo(photo_id):
    rows = get_rows()
    item = next((r for r in rows if r["id"] == photo_id), None)
    if not item:
        flash("Photo not found")
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        product_name = request.form.get("product_name", "").strip()
        photo_type = request.form.get("photo_type")
        file = request.files.get("photo")
        new_filename = None
        if file and file.filename:
            if not allowed_file(file.filename):
                flash("Only PNG, JPG, JPEG, WEBP images are allowed")
                return redirect(url_for("edit_photo", photo_id=photo_id))
            new_filename = datetime.now().strftime("%Y%m%d%H%M%S_") + secure_filename(file.filename)
            file.save(os.path.join(app.config["UPLOAD_FOLDER"], new_filename))
        update_row(photo_id, product_name, photo_type, new_filename)
        return redirect(url_for("dashboard"))
    return render_template("edit.html", item=item, photo_types=PHOTO_TYPES)


@app.route("/delete/<int:photo_id>", methods=["POST"])
@login_required
def delete_photo(photo_id):
    if session.get("role") != "admin":
        flash("Only admin can delete photos")
        return redirect(url_for("dashboard"))
    delete_row(photo_id)
    return redirect(url_for("dashboard"))


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


if __name__ == "__main__":
    init_excel()
    app.run(debug=True)
